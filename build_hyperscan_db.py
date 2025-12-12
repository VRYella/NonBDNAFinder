#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                   HYPERSCAN DATABASE BUILDER UTILITY                          ║
║         Build Hyperscan Pattern Databases for Fast Motif Detection           ║
╚══════════════════════════════════════════════════════════════════════════════╝

SCRIPT: build_hyperscan_db.py
AUTHOR: Dr. Venkata Rajesh Yella
VERSION: 2024.1
LICENSE: MIT

DESCRIPTION:
    Builds Hyperscan databases from Non-B DNA regex patterns. Can read patterns
    from TSV/Excel files (exported by export_patterns_full.py) or directly from
    motif_patterns.PATTERN_REGISTRY.
    
    Only Hyperscan-compatible patterns are included in the database. The utility
    automatically filters out incompatible patterns and reports them.
    
    The compiled database is serialized and saved to a .hsdb file that can be
    loaded quickly for high-performance pattern matching.

USAGE:
    # Build from all patterns in registry (exclude R-loop by default)
    python build_hyperscan_db.py --out all_patterns.hsdb
    
    # Build from specific detector class
    python build_hyperscan_db.py --class GQuadruplexDetector --out g4.hsdb
    
    # Build from exported TSV file
    python build_hyperscan_db.py --patterns patterns.tsv --out patterns.hsdb
    
    # Build from exported Excel file
    python build_hyperscan_db.py --patterns patterns.xlsx --out patterns.hsdb
    
    # Include R-loop patterns
    python build_hyperscan_db.py --out all_patterns.hsdb --include-rloop

OUTPUT FILES:
    - <out>.hsdb: Serialized Hyperscan database file
    
REQUIREMENTS:
    - hyperscan Python package (pip install hyperscan)
    - For reading Excel files: openpyxl (pip install openpyxl)
"""

import argparse
import csv
import sys
from pathlib import Path
from typing import List, Dict, Tuple, Any

try:
    import hyperscan
    HYPERSCAN_AVAILABLE = True
except ImportError:
    HYPERSCAN_AVAILABLE = False
    hyperscan = None

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from motif_patterns import PATTERN_REGISTRY
    PATTERN_REGISTRY_AVAILABLE = True
except ImportError:
    PATTERN_REGISTRY_AVAILABLE = False
    print("Warning: Could not import motif_patterns. Registry-based export disabled.", file=sys.stderr)


def test_hyperscan_compatible(pattern: str) -> bool:
    """
    Test if a pattern is compatible with Hyperscan.
    
    Args:
        pattern: Regex pattern string
        
    Returns:
        True if pattern compiles successfully, False otherwise
    """
    if not pattern:
        return False
    
    try:
        db = hyperscan.Database()
        db.compile(
            expressions=[pattern.encode('ascii')],
            ids=[0],
            elements=1
        )
        return True
    except Exception:
        return False


def load_patterns_from_tsv(tsv_path: Path) -> List[Dict[str, Any]]:
    """
    Load patterns from a TSV file exported by export_patterns_full.py.
    
    Args:
        tsv_path: Path to TSV file
        
    Returns:
        List of pattern dictionaries
    """
    patterns = []
    
    with open(tsv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter='\t')
        for row in reader:
            patterns.append(row)
    
    return patterns


def load_patterns_from_excel(excel_path: Path) -> List[Dict[str, Any]]:
    """
    Load patterns from an Excel file exported by export_patterns_full.py.
    
    Args:
        excel_path: Path to Excel file
        
    Returns:
        List of pattern dictionaries
    """
    if not EXCEL_AVAILABLE:
        print("Error: openpyxl not available. Cannot read Excel files.", file=sys.stderr)
        sys.exit(1)
    
    patterns = []
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Read data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        pattern_dict = dict(zip(headers, row))
        patterns.append(pattern_dict)
    
    return patterns


def load_patterns_from_registry(
    detector_class: str = None,
    include_rloop: bool = False
) -> List[Dict[str, Any]]:
    """
    Load patterns directly from PATTERN_REGISTRY.
    
    Args:
        detector_class: If specified, only load patterns from this detector
        include_rloop: If True, include R-loop patterns (default: False)
        
    Returns:
        List of pattern dictionaries
    """
    if not PATTERN_REGISTRY_AVAILABLE:
        print("Error: motif_patterns module not available.", file=sys.stderr)
        sys.exit(1)
    
    patterns = []
    
    for det_class, pattern_groups in PATTERN_REGISTRY.items():
        # Filter by detector class if specified
        if detector_class and det_class != detector_class:
            continue
        
        # Skip R-loop detector if not included
        if not include_rloop and det_class == 'RLoopDetector':
            continue
        
        for pattern_group, pattern_list in pattern_groups.items():
            for pattern_tuple in pattern_list:
                if len(pattern_tuple) < 9:
                    continue
                
                regex_pattern = pattern_tuple[0]
                pattern_id = pattern_tuple[1]
                name = pattern_tuple[2]
                subclass = pattern_tuple[3]
                
                # Skip empty patterns
                if not regex_pattern:
                    continue
                
                pattern_dict = {
                    'Class': det_class,
                    'Subclass': subclass,
                    'Pattern_Group': pattern_group,
                    'Pattern': regex_pattern,
                    'Pattern_ID': pattern_id,
                    'Name': name,
                }
                
                patterns.append(pattern_dict)
    
    return patterns


def filter_hyperscan_compatible(patterns: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Filter patterns to only Hyperscan-compatible ones.
    
    Args:
        patterns: List of pattern dictionaries
        
    Returns:
        Tuple of (compatible_patterns, incompatible_patterns)
    """
    compatible = []
    incompatible = []
    
    for pattern_dict in patterns:
        pattern = pattern_dict.get('Pattern', '')
        
        # Check if pattern was already marked as compatible
        hs_compat = pattern_dict.get('Hyperscan_compatible', '').lower()
        if hs_compat == 'yes':
            compatible.append(pattern_dict)
            continue
        elif hs_compat == 'no':
            incompatible.append(pattern_dict)
            continue
        
        # Test compatibility
        if test_hyperscan_compatible(pattern):
            compatible.append(pattern_dict)
        else:
            incompatible.append(pattern_dict)
    
    return compatible, incompatible


def build_hyperscan_database(patterns: List[Dict[str, Any]], output_path: Path) -> None:
    """
    Build and serialize a Hyperscan database from patterns.
    
    Args:
        patterns: List of pattern dictionaries (must be Hyperscan-compatible)
        output_path: Path to output .hsdb file
    """
    if not patterns:
        print("Error: No patterns to build database from.", file=sys.stderr)
        sys.exit(1)
    
    print(f"Building Hyperscan database with {len(patterns)} patterns...")
    
    # Prepare expressions and IDs
    expressions = []
    ids = []
    pattern_info = {}
    
    for idx, pattern_dict in enumerate(patterns):
        pattern = pattern_dict.get('Pattern', '')
        pattern_id = pattern_dict.get('Pattern_ID', f'PAT_{idx}')
        
        expressions.append(pattern.encode('ascii'))
        ids.append(idx)
        pattern_info[idx] = {
            'pattern_id': pattern_id,
            'name': pattern_dict.get('Name', ''),
            'pattern': pattern,
        }
    
    # Compile Hyperscan database
    try:
        db = hyperscan.Database()
        db.compile(
            expressions=expressions,
            ids=ids,
            elements=len(expressions),
            flags=[0] * len(expressions)  # Default flags
        )
        print(f"✓ Compiled {len(expressions)} patterns into Hyperscan database")
    except Exception as e:
        print(f"Error: Failed to compile Hyperscan database: {e}", file=sys.stderr)
        sys.exit(1)
    
    # Serialize database
    try:
        serialized = db.serialize()
        with open(output_path, 'wb') as f:
            f.write(serialized)
        print(f"✓ Serialized database to: {output_path}")
        print(f"  Database size: {len(serialized):,} bytes")
    except Exception as e:
        print(f"Error: Failed to serialize database: {e}", file=sys.stderr)
        sys.exit(1)


def main():
    """Main function for CLI."""
    # Check for Hyperscan availability early
    if not HYPERSCAN_AVAILABLE:
        print("Error: hyperscan package not available. Install with: pip install hyperscan", file=sys.stderr)
        print("This utility requires Hyperscan for building pattern databases.", file=sys.stderr)
        sys.exit(1)
    
    parser = argparse.ArgumentParser(
        description='Build Hyperscan databases from Non-B DNA regex patterns',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Build from all patterns in registry
  python build_hyperscan_db.py --out all_patterns.hsdb
  
  # Build from specific detector class
  python build_hyperscan_db.py --class GQuadruplexDetector --out g4.hsdb
  
  # Build from exported TSV file
  python build_hyperscan_db.py --patterns patterns.tsv --out patterns.hsdb
  
  # Build from exported Excel file
  python build_hyperscan_db.py --patterns patterns.xlsx --out patterns.hsdb
  
  # Include R-loop patterns
  python build_hyperscan_db.py --out all_patterns.hsdb --include-rloop

Output:
  Creates a serialized Hyperscan database file (.hsdb) for fast pattern matching.
  Only Hyperscan-compatible patterns are included.
        """
    )
    
    parser.add_argument(
        '--out',
        type=str,
        required=True,
        help='Output .hsdb file path'
    )
    
    parser.add_argument(
        '--patterns',
        type=str,
        help='Input patterns file (TSV or Excel). If not provided, loads from registry.'
    )
    
    parser.add_argument(
        '--class',
        type=str,
        dest='detector_class',
        help='Detector class name (e.g., GQuadruplexDetector). Only used with registry.'
    )
    
    parser.add_argument(
        '--include-rloop',
        action='store_true',
        help='Include R-loop patterns (excluded by default). Only used with registry.'
    )
    
    args = parser.parse_args()
    
    # Load patterns
    if args.patterns:
        patterns_file = Path(args.patterns)
        
        if not patterns_file.exists():
            print(f"Error: Patterns file not found: {patterns_file}", file=sys.stderr)
            sys.exit(1)
        
        print(f"Loading patterns from: {patterns_file}")
        
        if patterns_file.suffix == '.tsv':
            patterns = load_patterns_from_tsv(patterns_file)
        elif patterns_file.suffix in ['.xlsx', '.xls']:
            patterns = load_patterns_from_excel(patterns_file)
        else:
            print(f"Error: Unsupported file format: {patterns_file.suffix}", file=sys.stderr)
            print("Supported formats: .tsv, .xlsx", file=sys.stderr)
            sys.exit(1)
    else:
        print("Loading patterns from PATTERN_REGISTRY...")
        patterns = load_patterns_from_registry(
            detector_class=args.detector_class,
            include_rloop=args.include_rloop
        )
    
    print(f"Loaded {len(patterns)} patterns")
    
    # Filter to Hyperscan-compatible patterns
    print("Filtering Hyperscan-compatible patterns...")
    compatible, incompatible = filter_hyperscan_compatible(patterns)
    
    print(f"✓ Compatible patterns: {len(compatible)}")
    if incompatible:
        print(f"✗ Incompatible patterns: {len(incompatible)}")
        print("\nIncompatible patterns (will be excluded):")
        for pattern_dict in incompatible[:10]:  # Show first 10
            pattern_id = pattern_dict.get('Pattern_ID', 'N/A')
            name = pattern_dict.get('Name', 'N/A')
            print(f"  - {pattern_id}: {name}")
        if len(incompatible) > 10:
            print(f"  ... and {len(incompatible) - 10} more")
    
    if not compatible:
        print("Error: No Hyperscan-compatible patterns found.", file=sys.stderr)
        sys.exit(1)
    
    # Build database
    output_path = Path(args.out)
    if not output_path.suffix:
        output_path = output_path.with_suffix('.hsdb')
    
    build_hyperscan_database(compatible, output_path)
    
    print("\n✓ Hyperscan database built successfully!")


if __name__ == '__main__':
    main()
