#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║                    PATTERN EXPORT UTILITY                                     ║
║           Export Non-B DNA Patterns to TSV and Excel Formats                 ║
╚══════════════════════════════════════════════════════════════════════════════╝

SCRIPT: export_patterns_full.py
AUTHOR: Dr. Venkata Rajesh Yella
VERSION: 2024.1
LICENSE: MIT

DESCRIPTION:
    Exports all regex patterns from motif_patterns.PATTERN_REGISTRY to TSV 
    and Excel (.xlsx) formats with comprehensive metadata including:
    - Class (detector name)
    - Subclass (motif subclass)
    - Pattern_Group (pattern group within detector)
    - Pattern (the regex pattern)
    - Pattern_ID (unique identifier)
    - Name (human-readable name)
    - Min_Length (minimum length for matches)
    - Score_Type (scoring method)
    - Base_Score (base score value)
    - Python_re_compiles (whether pattern compiles in Python re)
    - Hyperscan_compatible (whether pattern is Hyperscan-compatible)
    - Description (motif description)
    - Reference (literature reference)

USAGE:
    # Export to both TSV and Excel (default, skip R-loop and empty patterns)
    python export_patterns_full.py --out patterns_export
    
    # Include R-loop patterns
    python export_patterns_full.py --out patterns_export --include-rloop
    
    # Include algorithmic/empty patterns
    python export_patterns_full.py --out patterns_export --include-empty
    
    # Export only TSV format
    python export_patterns_full.py --out patterns_export --format tsv
    
    # Export only Excel format
    python export_patterns_full.py --out patterns_export --format excel

OUTPUT FILES:
    - <out>.tsv: Tab-separated values file
    - <out>.xlsx: Excel workbook with formatted columns
"""

import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Dict, List, Tuple, Any

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.", file=sys.stderr)

try:
    from motif_patterns import PATTERN_REGISTRY
except ImportError:
    print("Error: Could not import motif_patterns. Make sure you're running from the correct directory.", file=sys.stderr)
    sys.exit(1)

# Try to import hyperscan for compatibility testing
try:
    import hyperscan
    HYPERSCAN_AVAILABLE = True
except ImportError:
    HYPERSCAN_AVAILABLE = False


def test_python_re_compile(pattern: str) -> bool:
    """
    Test if a pattern compiles with Python's re module.
    
    Args:
        pattern: Regex pattern string
        
    Returns:
        True if pattern compiles successfully, False otherwise
    """
    if not pattern:  # Empty patterns
        return False
    try:
        re.compile(pattern)
        return True
    except Exception:
        return False


def test_hyperscan_compatible(pattern: str) -> bool:
    """
    Test if a pattern is compatible with Hyperscan.
    
    Args:
        pattern: Regex pattern string
        
    Returns:
        True if pattern is Hyperscan-compatible, False otherwise
    """
    if not pattern:  # Empty patterns
        return False
    
    if not HYPERSCAN_AVAILABLE:
        # If Hyperscan is not available, do basic heuristic checks
        # Hyperscan doesn't support some features like backreferences, lookaheads, etc.
        unsupported_features = [
            r'\\[0-9]',  # Backreferences like \1, \2
            r'\(\?=',     # Positive lookahead
            r'\(\?!',     # Negative lookahead
            r'\(\?<=',    # Positive lookbehind
            r'\(\?<!',    # Negative lookbehind
            r'\(\?P<',    # Named groups
        ]
        for feature in unsupported_features:
            if re.search(feature, pattern):
                return False
        return True
    
    # If Hyperscan is available, try to compile it
    try:
        db = hyperscan.Database()
        db.compile(
            expressions=[pattern.encode('ascii')],
            ids=[0],
            elements=1
        )
        return True
    except Exception:
        return False


def extract_all_patterns(
    include_rloop: bool = False,
    include_empty: bool = False
) -> List[Dict[str, Any]]:
    """
    Extract all patterns from PATTERN_REGISTRY with metadata.
    
    Args:
        include_rloop: If True, include R-loop patterns (default: False)
        include_empty: If True, include algorithmic/empty patterns (default: False)
        
    Returns:
        List of dictionaries containing pattern metadata
    """
    patterns_data = []
    
    for detector_class, pattern_groups in PATTERN_REGISTRY.items():
        # Skip R-loop detector if not included
        if not include_rloop and detector_class == 'RLoopDetector':
            continue
        
        for pattern_group, pattern_list in pattern_groups.items():
            for pattern_tuple in pattern_list:
                # Pattern tuple structure:
                # (regex_pattern, pattern_id, name, subclass, min_length, 
                #  score_type, base_score, description, reference)
                
                if len(pattern_tuple) < 9:
                    continue  # Skip malformed patterns
                
                regex_pattern = pattern_tuple[0]
                pattern_id = pattern_tuple[1]
                name = pattern_tuple[2]
                subclass = pattern_tuple[3]
                min_length = pattern_tuple[4]
                score_type = pattern_tuple[5]
                base_score = pattern_tuple[6]
                description = pattern_tuple[7]
                reference = pattern_tuple[8]
                
                # Skip empty patterns unless explicitly included
                if not include_empty and not regex_pattern:
                    continue
                
                # Test pattern compilation
                py_compiles = test_python_re_compile(regex_pattern)
                hs_compatible = test_hyperscan_compatible(regex_pattern)
                
                pattern_data = {
                    'Class': detector_class,
                    'Subclass': subclass,
                    'Pattern_Group': pattern_group,
                    'Pattern': regex_pattern,
                    'Pattern_ID': pattern_id,
                    'Name': name,
                    'Min_Length': min_length,
                    'Score_Type': score_type,
                    'Base_Score': base_score,
                    'Python_re_compiles': 'Yes' if py_compiles else 'No',
                    'Hyperscan_compatible': 'Yes' if hs_compatible else 'No',
                    'Description': description,
                    'Reference': reference,
                }
                
                patterns_data.append(pattern_data)
    
    return patterns_data


def export_to_tsv(patterns_data: List[Dict[str, Any]], output_path: Path) -> None:
    """
    Export patterns to TSV format.
    
    Args:
        patterns_data: List of pattern dictionaries
        output_path: Path to output TSV file
    """
    if not patterns_data:
        print("Warning: No patterns to export.", file=sys.stderr)
        return
    
    fieldnames = [
        'Class', 'Subclass', 'Pattern_Group', 'Pattern', 'Pattern_ID',
        'Name', 'Min_Length', 'Score_Type', 'Base_Score',
        'Python_re_compiles', 'Hyperscan_compatible',
        'Description', 'Reference'
    ]
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter='\t')
        writer.writeheader()
        writer.writerows(patterns_data)
    
    print(f"✓ Exported {len(patterns_data)} patterns to TSV: {output_path}")


def export_to_excel(patterns_data: List[Dict[str, Any]], output_path: Path) -> None:
    """
    Export patterns to Excel (.xlsx) format with formatting.
    
    Args:
        patterns_data: List of pattern dictionaries
        output_path: Path to output Excel file
    """
    if not EXCEL_AVAILABLE:
        print("Error: openpyxl not available. Cannot export to Excel.", file=sys.stderr)
        return
    
    if not patterns_data:
        print("Warning: No patterns to export.", file=sys.stderr)
        return
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Non-B DNA Patterns"
    
    # Define headers
    headers = [
        'Class', 'Subclass', 'Pattern_Group', 'Pattern', 'Pattern_ID',
        'Name', 'Min_Length', 'Score_Type', 'Base_Score',
        'Python_re_compiles', 'Hyperscan_compatible',
        'Description', 'Reference'
    ]
    
    # Style header row
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_idx, pattern_data in enumerate(patterns_data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            value = pattern_data.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=False)
    
    # Adjust column widths
    column_widths = {
        'Class': 20,
        'Subclass': 20,
        'Pattern_Group': 25,
        'Pattern': 50,
        'Pattern_ID': 15,
        'Name': 30,
        'Min_Length': 12,
        'Score_Type': 20,
        'Base_Score': 12,
        'Python_re_compiles': 18,
        'Hyperscan_compatible': 20,
        'Description': 40,
        'Reference': 20,
    }
    
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = column_widths.get(header, 15)
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    wb.save(output_path)
    print(f"✓ Exported {len(patterns_data)} patterns to Excel: {output_path}")


def main():
    """Main function for CLI."""
    parser = argparse.ArgumentParser(
        description='Export Non-B DNA regex patterns to TSV and Excel formats',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Export to both TSV and Excel (default)
  python export_patterns_full.py --out patterns_export
  
  # Include R-loop patterns
  python export_patterns_full.py --out patterns_export --include-rloop
  
  # Include algorithmic/empty patterns
  python export_patterns_full.py --out patterns_export --include-empty
  
  # Export only TSV format
  python export_patterns_full.py --out patterns_export --format tsv
  
  # Export only Excel format
  python export_patterns_full.py --out patterns_export --format excel

Output:
  Creates <out>.tsv and/or <out>.xlsx with columns:
  Class, Subclass, Pattern_Group, Pattern, Pattern_ID, Name, Min_Length,
  Score_Type, Base_Score, Python_re_compiles, Hyperscan_compatible,
  Description, Reference
        """
    )
    
    parser.add_argument(
        '--out',
        type=str,
        required=True,
        help='Output file path (without extension)'
    )
    
    parser.add_argument(
        '--format',
        type=str,
        choices=['both', 'tsv', 'excel'],
        default='both',
        help='Export format: both (default), tsv, or excel'
    )
    
    parser.add_argument(
        '--include-rloop',
        action='store_true',
        help='Include R-loop patterns (excluded by default)'
    )
    
    parser.add_argument(
        '--include-empty',
        action='store_true',
        help='Include algorithmic/empty patterns (excluded by default)'
    )
    
    args = parser.parse_args()
    
    # Extract patterns
    print("Extracting patterns from PATTERN_REGISTRY...")
    patterns_data = extract_all_patterns(
        include_rloop=args.include_rloop,
        include_empty=args.include_empty
    )
    
    if not patterns_data:
        print("Error: No patterns found to export.", file=sys.stderr)
        sys.exit(1)
    
    print(f"Found {len(patterns_data)} patterns to export")
    
    # Export based on format
    output_base = Path(args.out)
    
    if args.format in ['both', 'tsv']:
        tsv_path = output_base.with_suffix('.tsv')
        export_to_tsv(patterns_data, tsv_path)
    
    if args.format in ['both', 'excel']:
        if not EXCEL_AVAILABLE:
            print("Warning: Skipping Excel export (openpyxl not available)", file=sys.stderr)
        else:
            excel_path = output_base.with_suffix('.xlsx')
            export_to_excel(patterns_data, excel_path)
    
    print("\n✓ Pattern export completed successfully!")


if __name__ == '__main__':
    main()
