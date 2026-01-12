#!/usr/bin/env python3
"""
Publication Export Module
=========================

Generates unified Excel workbooks with comprehensive analysis results
suitable for publication and manuscript preparation.

Features:
- Summary sheet with all genomes
- Per-genome detailed sheets
- Comparative analysis sheet
- Statistical tables
- Publication-ready formatting

Usage:
    python publication_export.py [--input-dir INPUT_DIR] [--comparative-dir COMPARATIVE_DIR] [--output OUTPUT]
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Add current directory to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def create_summary_sheet(wb: Workbook, batch_summary: Dict, comparative_results: Dict):
    """
    Create summary sheet with overview of all genomes.
    
    Args:
        wb: Workbook object
        batch_summary: Batch analysis summary
        comparative_results: Comparative analysis results
    """
    ws = wb.active
    ws.title = "Summary"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Title
    ws['A1'] = "NonBDNAFinder - Batch Genome Analysis Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    
    # Metadata
    ws['A3'] = "Analysis Date:"
    ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws['A4'] = "Total Genomes:"
    ws['B4'] = comparative_results['summary']['total_genomes']
    ws['A5'] = "Total Motifs:"
    ws['B5'] = comparative_results['summary']['total_motifs_detected']
    ws['A6'] = "Total Sequence Length:"
    ws['B6'] = f"{comparative_results['summary']['total_sequence_length']:,} bp"
    
    # Per-genome summary table
    ws['A8'] = "Genome Name"
    ws['B8'] = "Length (bp)"
    ws['C8'] = "Total Motifs"
    ws['D8'] = "Density (motifs/kb)"
    ws['E8'] = "Coverage (%)"
    ws['F8'] = "Unique Classes"
    ws['G8'] = "Unique Subclasses"
    ws['H8'] = "Class Diversity"
    
    # Apply header styling
    for col in range(1, 9):
        cell = ws.cell(row=8, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Fill genome data
    row = 9
    for genome_name, stats in comparative_results['genome_statistics'].items():
        ws.cell(row=row, column=1, value=genome_name)
        ws.cell(row=row, column=2, value=stats['total_length'])
        ws.cell(row=row, column=3, value=stats['total_motifs'])
        ws.cell(row=row, column=4, value=round(stats['density_per_kb'], 2))
        ws.cell(row=row, column=5, value=round(stats['coverage_percent'], 2))
        ws.cell(row=row, column=6, value=stats['unique_classes'])
        ws.cell(row=row, column=7, value=stats['unique_subclasses'])
        ws.cell(row=row, column=8, value=round(stats['class_diversity'], 3))
        row += 1
    
    # Auto-size columns
    from openpyxl.cell.cell import MergedCell
    for column in ws.columns:
        max_length = 0
        column_letter = None
        for cell in column:
            # Skip merged cells
            if isinstance(cell, MergedCell):
                continue
            if column_letter is None:
                column_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if column_letter:
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def create_genome_sheets(wb: Workbook, batch_results: Dict, comparative_results: Dict):
    """
    Create detailed sheets for each genome.
    
    Args:
        wb: Workbook object
        batch_results: Batch analysis results
        comparative_results: Comparative analysis results
    """
    for genome_name, genome_data in batch_results.items():
        # Create sanitized sheet name (max 31 chars for Excel)
        sheet_name = genome_name.replace('.fna', '').replace('.fasta', '').replace(' ', '_')[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Header
        ws['A1'] = f"Detailed Results: {genome_name}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:F1')
        
        # Statistics
        stats = comparative_results['genome_statistics'][genome_name]
        ws['A3'] = "Statistics"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = "Total Length:"
        ws['B4'] = f"{stats['total_length']:,} bp"
        ws['A5'] = "Total Motifs:"
        ws['B5'] = stats['total_motifs']
        ws['A6'] = "Density:"
        ws['B6'] = f"{stats['density_per_kb']:.2f} motifs/kb"
        ws['A7'] = "Coverage:"
        ws['B7'] = f"{stats['coverage_percent']:.2f}%"
        
        # Class distribution
        ws['A9'] = "Motif Class Distribution"
        ws['A9'].font = Font(bold=True)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        ws['A10'] = "Class"
        ws['B10'] = "Count"
        ws['C10'] = "Percentage"
        
        for col in ['A10', 'B10', 'C10']:
            ws[col].fill = header_fill
            ws[col].font = header_font
        
        row = 11
        total_motifs = stats['total_motifs']
        for class_name, count in sorted(stats['class_counts'].items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1, value=class_name)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/total_motifs*100:.2f}%")
            row += 1


def create_comparative_sheet(wb: Workbook, comparative_results: Dict):
    """
    Create comparative analysis sheet.
    
    Args:
        wb: Workbook object
        comparative_results: Comparative analysis results
    """
    ws = wb.create_sheet(title="Comparative Analysis")
    
    # Header
    ws['A1'] = "Cross-Genome Comparative Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Summary statistics
    summary = comparative_results['summary']
    ws['A3'] = "Summary Statistics"
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A4'] = "Average Density (motifs/kb):"
    ws['B4'] = round(summary['average_density_per_kb'], 2)
    ws['A5'] = "Average Coverage (%):"
    ws['B5'] = round(summary['average_coverage_percent'], 2)
    ws['A6'] = "Average Class Diversity:"
    ws['B6'] = round(summary['average_class_diversity'], 3)
    
    # Notable patterns
    ws['A8'] = "Notable Patterns"
    ws['A8'].font = Font(bold=True, size=12)
    
    ws['A9'] = "Highest Density Genome:"
    ws['B9'] = summary['highest_density_genome']['name']
    ws['C9'] = f"{summary['highest_density_genome']['density']:.2f} motifs/kb"
    
    ws['A10'] = "Lowest Density Genome:"
    ws['B10'] = summary['lowest_density_genome']['name']
    ws['C10'] = f"{summary['lowest_density_genome']['density']:.2f} motifs/kb"
    
    ws['A11'] = "Highest Diversity Genome:"
    ws['B11'] = summary['highest_diversity_genome']['name']
    ws['C11'] = f"H={summary['highest_diversity_genome']['diversity']:.3f}"
    
    ws['A12'] = "Lowest Diversity Genome:"
    ws['B12'] = summary['lowest_diversity_genome']['name']
    ws['C12'] = f"H={summary['lowest_diversity_genome']['diversity']:.3f}"
    
    # Class frequency comparison
    ws['A14'] = "Motif Class Frequency Comparison"
    ws['A14'].font = Font(bold=True, size=12)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Create header row with genome names
    ws['A15'] = "Motif Class"
    ws['A15'].fill = header_fill
    ws['A15'].font = header_font
    
    genome_names = list(comparative_results['genome_statistics'].keys())
    for idx, genome_name in enumerate(genome_names, start=2):
        cell = ws.cell(row=15, column=idx)
        cell.value = genome_name
        cell.fill = header_fill
        cell.font = header_font
    
    # Fill frequency data
    row = 16
    for class_name, frequencies in comparative_results['class_frequencies'].items():
        ws.cell(row=row, column=1, value=class_name)
        
        for idx, freq_data in enumerate(frequencies, start=2):
            ws.cell(row=row, column=idx, value=round(freq_data['frequency'] * 100, 2))
        
        row += 1


def create_statistics_sheet(wb: Workbook, comparative_results: Dict):
    """
    Create statistical analysis sheet.
    
    Args:
        wb: Workbook object
        comparative_results: Comparative analysis results
    """
    ws = wb.create_sheet(title="Statistical Analysis")
    
    # Header
    ws['A1'] = "Statistical Analysis Results"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    # Enrichment patterns
    ws['A3'] = "Enrichment/Depletion Patterns"
    ws['A3'].font = Font(bold=True, size=12)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A4'] = "Motif Class"
    ws['B4'] = "Mean Frequency"
    ws['C4'] = "Std Dev"
    ws['D4'] = "Coeff. Variation"
    ws['E4'] = "Enriched Genomes"
    ws['F4'] = "Depleted Genomes"
    
    for col in ['A4', 'B4', 'C4', 'D4', 'E4', 'F4']:
        ws[col].fill = header_fill
        ws[col].font = header_font
    
    row = 5
    for class_name, enrichment in comparative_results['enrichment_patterns'].items():
        ws.cell(row=row, column=1, value=class_name)
        ws.cell(row=row, column=2, value=round(enrichment['mean_frequency'], 4))
        ws.cell(row=row, column=3, value=round(enrichment['std_frequency'], 4))
        ws.cell(row=row, column=4, value=round(enrichment['coefficient_variation'], 3))
        
        # Enriched genomes
        enriched = ', '.join([g['genome'] for g in enrichment['enriched_genomes']])
        ws.cell(row=row, column=5, value=enriched if enriched else "None")
        
        # Depleted genomes
        depleted = ', '.join([g['genome'] for g in enrichment['depleted_genomes']])
        ws.cell(row=row, column=6, value=depleted if depleted else "None")
        
        row += 1
    
    # Correlations
    if comparative_results['correlations']:
        ws[f'A{row+2}'] = "Feature Correlations"
        ws[f'A{row+2}'].font = Font(bold=True, size=12)
        
        ws[f'A{row+3}'] = "Feature Pair"
        ws[f'B{row+3}'] = "Correlation"
        ws[f'C{row+3}'] = "P-value"
        ws[f'D{row+3}'] = "Significant"
        
        for col in [f'A{row+3}', f'B{row+3}', f'C{row+3}', f'D{row+3}']:
            ws[col].fill = header_fill
            ws[col].font = header_font
        
        curr_row = row + 4
        for pair_name, corr_data in comparative_results['correlations'].items():
            ws.cell(row=curr_row, column=1, value=pair_name.replace('_', ' ').title())
            ws.cell(row=curr_row, column=2, value=round(corr_data['correlation'], 3))
            ws.cell(row=curr_row, column=3, value=round(corr_data['p_value'], 4))
            ws.cell(row=curr_row, column=4, value="Yes" if corr_data['significant'] else "No")
            curr_row += 1


def generate_publication_excel(batch_dir: str, comparative_dir: str, output_file: str):
    """
    Generate unified publication Excel workbook.
    
    Args:
        batch_dir: Directory with batch results
        comparative_dir: Directory with comparative results
        output_file: Output Excel file path
    """
    print("Loading results...")
    
    # Load batch summary
    with open(os.path.join(batch_dir, 'batch_summary.json'), 'r') as f:
        batch_summary = json.load(f)
    
    # Load comparative results
    with open(os.path.join(comparative_dir, 'comparative_analysis.json'), 'r') as f:
        comparative_results = json.load(f)
    
    # Load individual genome results
    batch_results = {}
    for genome_info in batch_summary['genomes']:
        genome_file = genome_info['genome_file']
        safe_name = genome_file.replace('.fna', '').replace('.fasta', '').replace(' ', '_')
        motifs_file = os.path.join(batch_dir, f"{safe_name}_motifs.json")
        
        if os.path.exists(motifs_file):
            with open(motifs_file, 'r') as f:
                motifs = json.load(f)
            
            batch_results[genome_file] = {
                'info': genome_info,
                'motifs': motifs
            }
    
    print("Creating Excel workbook...")
    wb = Workbook()
    
    # Create sheets
    print("  → Summary sheet")
    create_summary_sheet(wb, batch_summary, comparative_results)
    
    print("  → Per-genome sheets")
    create_genome_sheets(wb, batch_results, comparative_results)
    
    print("  → Comparative analysis sheet")
    create_comparative_sheet(wb, comparative_results)
    
    print("  → Statistical analysis sheet")
    create_statistics_sheet(wb, comparative_results)
    
    # Save workbook
    print(f"Saving workbook: {output_file}")
    wb.save(output_file)
    print(f"✓ Excel workbook saved successfully")


def main():
    """Main function for publication export"""
    parser = argparse.ArgumentParser(
        description='Generate unified publication Excel workbook'
    )
    parser.add_argument(
        '--batch-dir',
        default='batch_results',
        help='Batch results directory (default: batch_results)'
    )
    parser.add_argument(
        '--comparative-dir',
        default='comparative_results',
        help='Comparative results directory (default: comparative_results)'
    )
    parser.add_argument(
        '--output',
        default='publication_results.xlsx',
        help='Output Excel file (default: publication_results.xlsx)'
    )
    
    args = parser.parse_args()
    
    print("="*80)
    print("PUBLICATION EXCEL EXPORT")
    print("="*80)
    print(f"Batch directory: {args.batch_dir}")
    print(f"Comparative directory: {args.comparative_dir}")
    print(f"Output file: {args.output}\n")
    
    try:
        generate_publication_excel(args.batch_dir, args.comparative_dir, args.output)
        
        print("\n" + "="*80)
        print("EXPORT COMPLETE")
        print("="*80)
        print(f"Publication Excel workbook: {args.output}")
        print("\nWorkbook contains:")
        print("  • Summary sheet with all genomes")
        print("  • Per-genome detailed sheets")
        print("  • Comparative analysis sheet")
        print("  • Statistical analysis sheet")
        print("="*80)
        
        return 0
        
    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    sys.exit(main())
